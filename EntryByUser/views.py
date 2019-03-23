from django.contrib.auth.decorators import login_required
from django.shortcuts import render
#  Create your views here.
from openpyxl import Workbook
from datetime import datetime
from .models import DataEntry
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.http import HttpResponse,HttpResponseRedirect
from .forms import DailyEntryForm
from openpyxl.utils import get_column_letter
import xlwt

from django.contrib.auth.models import User

def export_users_xls(request,fromdate,todate):



    rows = DataEntry.objects.all().filter(date__range=[fromdate,todate])#["2019-01-01", "2019-03-22"])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Data'


    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Define the titles for columns
    columns = ['Date', 'Name', 'Town', 'Area','TotalCalls','PC','Secondary Order Value',
    'Secondary Collection Value','CumSecondarySales','CumSecondaryCollectionValue'
       ,'GTOnly','HAPromoterStores','MTStores','CummCollection'
       ,'NewStoresToday','NewStoresComm','Remakrs']

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal='center')
        cell.value = column_title

    # Iterate through all movies
    for DataRow in rows:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            DataRow.date,
            DataRow.Name,
            DataRow.Town,
            DataRow.Area,
            DataRow.TotalCalls,
            DataRow.PC,
            DataRow.SecondaryOrderValue,
            DataRow.SecondaryCollectionValue,
            DataRow.CumSecondarySales,
            DataRow.CumSecondaryCollectionValue,
            DataRow.GTOnly,
            DataRow.HAPromoterStores,
            DataRow.MTStores,
            DataRow.CummCollection,
            DataRow.NewStoresToday,
            DataRow.NewStoresComm,
            DataRow.remarks,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):

            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center')
            cell.value = cell_value

    workbook.save(response)

    return response


def DataEntryView(request):

     # if this is a POST request we need to process the form data
    if request.method == 'POST':
        form = DailyEntryForm(request.POST)
        # check whether it's valid:
        if form.is_valid():
            # process the data in form.cleaned_data as required
            # ...

            empid = form.cleaned_data['EmpId']

            # redirect to a new URL:
            form.save()
           # url = reverse('Data', kwargs={'Emp':form.Emp})
            return HttpResponseRedirect('Data/'+ str(empid) ,)





    # if a GET (or any other method) we'll create a blank form
    else:
        form = DailyEntryForm()

    return render(request, 'DailyEntry.html',{'form' : form})

def ViewData(request,Emp):
     data = DataEntry.objects.filter(EmpId=Emp).order_by('-id')[:1]

     ShowData = {
        "Data": data
     }
     return render(request,"DailyShow.html", ShowData)


@login_required(login_url='/login')
def AdminView(request):
    data = DataEntry.objects.filter(date__range=["2019-01-01", "2019-03-22"])
    ShowData = {
        "Data": data
     }
    return render(request,"AdminView.html", ShowData)
